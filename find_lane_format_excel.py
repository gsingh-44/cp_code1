import argparse
import pandas as pd
import os
#set my me to avoid error 15 for ocr
os.environ['KMP_DUPLICATE_LIB_OK']='True'
import easyocr
import cv2
import platform
import sys
from pathlib import Path
from PIL import Image
#import xlswriter
#import xlsxwriter
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill

import torch  
fill_green = PatternFill(patternType='solid', fgColor='BAFFC9')
fill_red = PatternFill(patternType='solid', fgColor='FFB3BA')
fill_gray = PatternFill(patternType='solid', fgColor='999DA0')

def highlight_rows_left(row):
    value = row.loc['X1']
    value2 = row.loc['X2']
    ocr_text = row.loc['OCR_text']
    #print(ocr_text)
    

    
    
    if value > 725 and value2 <1400:
        color = '#BAFFC9' # Green
    else:
        color = '#FFB3BA' # Red
    return ['background-color: {}'.format(color) for r in row]

def highlight_rows_mid(row):
    value = row.loc['X1']
    value2 = row.loc['X2']
    ocr_text = row.loc['OCR_text']
    #print(ocr_text)
    

    
    
    if value > 1450 and value2 <1870:
        color = '#BAFFC9' # Green
    else:
        color = '#FFB3BA' # Grey
    return ['background-color: {}'.format(color) for r in row]   

def highlight_rows_right(row):
    value = row.loc['X1']
    value2 = row.loc['X2']
    ocr_text = row.loc['OCR_text']
    #print(ocr_text)
    

    
    
    if value > 1900 and value2 <2530:
        color = '#BAFFC9' # Red
    else:
        color = '#FFB3BA' # Green
    
    return ['background-color: {}'.format(color) for r in row]    


def till_next_character_match(detected_string_index, given_string_index, detected_string, given_string, length):
    stored_given_string_index = given_string_index
    while detected_string_index < len(detected_string):
        given_string_index = stored_given_string_index
        while given_string_index < len(given_string):
            if given_string_index < len(given_string) and detected_string_index < len(detected_string) and given_string[given_string_index] == detected_string[detected_string_index]:
                detected_string_index = detected_string_index + 1
                given_string_index = given_string_index + 1
                length = length + 1
                return detected_string_index, given_string_index,length
            given_string_index = given_string_index + 1
        detected_string_index = detected_string_index + 1
    return detected_string_index, given_string_index, length

def find_max_matched(detected_string, given_strings):
    count_map = {}
    for given_string in given_strings:
        major_index = 0
        max_len = 0
        while major_index < len(detected_string):
            detected_string_index = major_index
            given_string_index = 0
            length = 0
            while detected_string_index < len(detected_string):
                detected_string_index, given_string_index,length  = till_next_character_match(detected_string_index, given_string_index, detected_string, given_string, length)
            max_len = max(max_len, length)
            count_map[given_string] = max_len
            major_index = major_index +1
    print(count_map)
    max_matched_string = ""
    max_matched_string_count = 0
    for key in count_map:
        if count_map[key] > max_matched_string_count:
            max_matched_string_count = count_map[key]
            
            max_matched_string = key
             
            
#     print(max_matched_string)
    return max_matched_string, max_matched_string_count
i = 1


map = pd.read_excel('maping.xlsx')
inf = pd.read_excel('rightlane.xlsx')
wb = openpyxl.load_workbook("rightlane.xlsx")
ws = wb['rightlane']

for count in range(len(inf)):
    ocr_text = str(inf.loc[count, 'OCR_text'])
    #map_df = pd.DataFrame(map)
    print([str(map.loc[index, 'left_lane']), str(map.loc[index, 'mid_lane']), str(map.loc[index, 'right_lane'])])
    match_count_old = 0
    internal_counter = 0
#print(len(map))
    for index in range (len(map)):
        internal_counter += 1  
        #print("for loop", internal_counter)
        #print('value=', map.loc[index, 'left_lane'])
        max_matched_string_new, match_count = find_max_matched(ocr_text, [str(map.loc[index, 'left_lane']), str(map.loc[index, 'mid_lane']), str(map.loc[index, 'right_lane'])])
        if match_count > match_count_old:
          max_matched_string = max_matched_string_new
          match_count_old = match_count

 #print(map.all)
    #print(max_matched_string, match_count_old)
    #print(type(max_matched_string))
    #print(type(ocr_text))
    ex_in_lane = 0
    
    if match_count_old >= 3:
        i, j = np.where(map.values == str(max_matched_string))
        #print(max_matched_string in map['left_lane'].unique())
        if(max_matched_string in map['left_lane'].unique()) is True:
            ex_in_lane=1
        elif(max_matched_string in map['mid_lane'].unique()) is True:
          ex_in_lane=2
        elif(max_matched_string in map['right_lane'].unique()) is True:
            ex_in_lane=3

            #print("not ex", max_matched_string) 
 #print(max_matched_string in map[''].values)
    else:
            ws.cell(count+2,7).value = str("containr") +str( " ") + str(ocr_text) + str(" ") + str("not scanned at the sorter")
            ws.cell(count+2, 1).fill = fill_gray
            ws.cell(count+2, 2).fill = fill_gray
            ws.cell(count+2, 3).fill = fill_gray
            ws.cell(count+2, 4).fill = fill_gray
            ws.cell(count+2, 5).fill = fill_gray
            ws.cell(count+2, 6).fill = fill_gray
            ws.cell(count+2, 7).fill = fill_gray
    print('ex in lane=', ex_in_lane)
    #print("count=" , count)
    value = ws.cell(count+2, 2).value
    value2 = ws.cell(count+2, 4).value
    if value > 725 and value2 <1400:
        found_in_ln = 1
    elif  value > 1450 and value2 <1919:
        found_in_ln = 2
    elif value > 1920 and value2 <2550:
         found_in_ln = 3
    else:
        ws.cell(count+2, 7).value = str("Found outside the range")     

    if found_in_ln == ex_in_lane:


    if ex_in_lane ==1:
         
         #print(value, value2)
         if value > 725 and value2 <1400:
            ws.cell(count+2, 1).fill = fill_green
            ws.cell(count+2, 2).fill = fill_green
            ws.cell(count+2, 3).fill = fill_green
            ws.cell(count+2, 4).fill = fill_green
            ws.cell(count+2, 5).fill = fill_green
            ws.cell(count+2, 6).fill = fill_green
            ws.cell(count+2, 7).value = str("Desitnation TILSONRBURG")
            
           #color = '#BAFFC9' # Green
         else:
            ws.cell(count+2, 1).fill = fill_red
            ws.cell(count+2, 2).fill = fill_red
            ws.cell(count+2, 3).fill = fill_red
            ws.cell(count+2, 4).fill = fill_red
            ws.cell(count+2, 5).fill = fill_red
            ws.cell(count+2, 6).fill = fill_red
            #ws['A.count'].fill = fill_red
           #color = '#FFB3BA' # Red

      #dt = inf.style.apply(highlight_rows_left, axis=1) 
    if ex_in_lane ==2:
        # value = ws.cell(count+2, 2).value
        # value2 = ws.cell(count+2, 4).value
         #print(value, value2)
         if value > 1450 and value2 <1919:
            ws.cell(count+2, 1).fill = fill_green
            ws.cell(count+2, 2).fill = fill_green
            ws.cell(count+2, 3).fill = fill_green
            ws.cell(count+2, 4).fill = fill_green
            ws.cell(count+2, 5).fill = fill_green
            ws.cell(count+2, 6).fill = fill_green
            ws.cell(count+2, 7).value = str("Destination SARNIA")
           #color = '#BAFFC9' # Green
         else:
           #color = '#FFB3BA' # Red
            ws.cell(count+2, 1).fill = fill_red
            ws.cell(count+2, 2).fill = fill_red
            ws.cell(count+2, 3).fill = fill_red
            ws.cell(count+2, 4).fill = fill_red
            ws.cell(count+2, 5).fill = fill_red
            ws.cell(count+2, 6).fill = fill_red
      #dt = inf.style.apply(highlight_rows_mid, axis=1) 
    if ex_in_lane ==3:
         #value = ws.cell(count+2, 2).value
         #value2 = ws.cell(count+2, 4).value
         #print(value, value2)
         if value > 1920 and value2 <2550:
            ws.cell(count+2, 1).fill = fill_green
            ws.cell(count+2, 2).fill = fill_green
            ws.cell(count+2, 3).fill = fill_green
            ws.cell(count+2, 4).fill = fill_green
            ws.cell(count+2, 5).fill = fill_green
            ws.cell(count+2, 6).fill = fill_green
            ws.cell(count+2, 7).value = str("Desitnation ST THOMAS")
           #color = '#BAFFC9' # Green
         else:
            ws.cell(count+2, 1).fill = fill_red
            ws.cell(count+2, 2).fill = fill_red
            ws.cell(count+2, 3).fill = fill_red
            ws.cell(count+2, 4).fill = fill_red
            ws.cell(count+2, 5).fill = fill_red
            ws.cell(count+2, 6).fill = fill_red
           #color = '#FFB3BA' # Red
      #dt = inf.style.apply(highlight_rows_right, axis=1)      
    #dt.to_excel('file3.xlsx', index=False ) 
    # 
wb.save('neew.xlsx')